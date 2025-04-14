import time
import unittest

import ddddocr
import docx
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

import utils

MAX_TRIES = 4

wb = openpyxl.Workbook()
# 创建一个sheet并加上名称和所在位置，第一个位置索引号是0
wb.create_sheet("广西农作物病虫疫情信息调度指挥平台", 0)
sheet = wb["广西农作物病虫疫情信息调度指挥平台"]
# 写入表头
headers = ["页面", "检测结果"]
sheet.append(headers)

# 创建Word文档对象
doc = docx.Document()


class PPMSGX():

    def __init__(self):
        # self.cookies = {}
        self.init_chrome()

    def init_chrome(self):
        self.driver = utils.new_chrome()
        self.driver.maximize_window()

    def re_connet(self):
        self.driver.quit()
        self.driver = utils.new_chrome()
        self.driver.maximize_window()
        print("系统重启")

    @utils.retry(MAX_TRIES)
    def test_guangXi(self):
        username = input("广西省级系统生产环境巡检开始\n请输入登录用户名：")
        password = input("请输入登录密码：")
        self.driver.maximize_window()
        # self.driver.get(base_url + "login")
        self.driver.get("https://gx.pestiot.com/login")
        self.driver.implicitly_wait(4)

        # 登录
        self.driver.find_element(By.XPATH, '//input[@placeholder="请输入用户名"]').send_keys(username)
        self.driver.find_element(By.XPATH, '//input[@name="password"]').send_keys(password)
        yanzhengma_image = self.driver.find_element(By.XPATH, '//img[contains(@src,"data:image/png;base64")]')
        img_bytes = yanzhengma_image.screenshot_as_png
        yzm = ddddocr.DdddOcr(show_ad=False).classification(img_bytes)
        self.driver.find_element(By.XPATH, '//input[@placeholder="请输入验证码"]').send_keys(yzm)
        time.sleep(3)
        self.driver.find_element(By.XPATH, '//button[@type="button"]').click()

        # 首页全年任务数量元素校验
        try:
            # 等待页面加载完成
            # WebDriverWait(self.driver, 15).until(lambda d: d.execute_script('return document.readyState') == 'complete')
            time.sleep(12)
            # 获取页面总高度和宽度
            total_height = self.driver.execute_script("return document.body.parentElement.scrollHeight")
            total_width = self.driver.execute_script("return document.body.parentElement.scrollWidth")
            # 调整窗口尺寸
            self.driver.set_window_size(total_width, total_height - 280)
            element = WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, '//div[@class="legendBox"]')))
            unittest.TestCase.assertTrue(element is not None, "登录成功，成功打开首页，且指定元素存在")
            utils.g_logger.info(f"登录成功，成功打开首页")
            sheet.append(["首页", "正常"])

            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/首页.png", doc, "首页")

        except Exception as e:
            utils.g_logger.info("登录成功，首页显示异常")
            sheet.append(["首页", "异常"])
            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/首页.png", doc, "首页")

        # 数据填报工作平台页
        self.driver.find_element(By.XPATH, "//div[starts-with(@class,'navi-item')][text()='数据填报 ']").click()
        try:
            self.driver.find_element(By.XPATH,
                                     "//li[contains(@class,'el-menu-item first-menu')]//span[text()='工作平台']").click()
            time.sleep(8)
            element = WebDriverWait(self.driver, 15).until(EC.visibility_of_element_located(
                (By.XPATH, '//div[@class="legendBox"]')))
            unittest.TestCase.assertTrue(element is not None, "成功打开数据填报工作平台，且指定元素存在")
            utils.g_logger.info("数据填报-工作平台页显示正常")
            sheet.append(["数据填报-工作平台", "正常"])
            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_工作平台.png", doc, "数据填报_工作平台")
        except Exception as e:
            utils.g_logger.info("数据填报-工作平台页显示异常")
            sheet.append(["数据填报-工作平台", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_工作平台.png", doc, "数据填报_工作平台")

        # 数据填报任务填报页
        self.driver.find_element(By.XPATH, "//li[@class='el-menu-item first-menu']//span[text()='数据填报']").click()

        try:
            time.sleep(10)
            self.driver.find_element(By.XPATH, "//i[@class='el-icon-arrow-down']").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//li[starts-with(@id,'reportTree')][1]")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开任务填报页，且指定元素存在")
            utils.g_logger.info("数据填报-任务填报页显示正常")
            sheet.append(["数据填报-任务填报", "正常"])
            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_任务填报.png", doc, "数据填报_任务填报")
        except Exception as e:
            utils.g_logger.info("数据填报-任务填报页显示异常")
            sheet.append(["数据填报-任务填报", "异常"])
            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_任务填报.png", doc, "数据填报_任务填报")

        # 数据填报数据查询页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='数据查询']").click()
        try:
            time.sleep(10)
            self.driver.find_element(By.XPATH, "//div[@id='reportboxTree']//i[@class='el-icon-arrow-down']").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//li[starts-with(@id,'reportTree')][1]")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开数据填报-数据查询页，且报表列表存在")
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开数据填报-数据查询页，且站点列表存在")
            utils.g_logger.info("数据填报-数据查询页显示正常")
            sheet.append(["数据填报-数据查询", "正常"])
            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_数据查询.png", doc, "数据填报_数据查询")
            doc.save('outputs/广西截图.docx')
        except Exception as e:
            utils.g_logger.info("数据填报-数据查询页显示异常")
            sheet.append(["数据填报-数据查询", "异常"])
            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_数据查询.png", doc, "数据填报_数据查询")

        # 数据填报数据汇总页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='数据汇总']").click()
        try:
            time.sleep(12)
            self.driver.find_element(By.XPATH, "//div[@id='reportboxTree']//i[@class='el-icon-arrow-down']").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//li[starts-with(@id,'reportTree')][1]")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开数据汇总页，且报表列表存在")
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开数据汇总页，且站点列表存在")
            utils.g_logger.info("数据填报-数据汇总页显示正常")
            sheet.append(["数据填报-数据汇总", "正常"])
            # 获取网页截图并保存至word文档
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_数据汇总.png", doc, "数据填报_数据汇总")

        except Exception as e:
            utils.g_logger.info("数据填报-数据汇总页显示异常")
            sheet.append(["数据填报-数据汇总", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_数据汇总.png", doc, "数据填报_数据汇总")

        # 数据填报催报查询页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='催报查询']").click()
        try:
            time.sleep(12)
            self.driver.find_element(By.XPATH, "//div[@id='reportboxTree']//i[@class='el-icon-arrow-down']").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//li[starts-with(@id,'reportTree')][1]")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开催报查询页，且报表列表存在")
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开催报查询页，且站点列表存在")
            utils.g_logger.info("数据填报-催报查询页显示正常")
            sheet.append(["数据填报-催报查询", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_催报查询.png", doc, "数据填报_催报查询")
        except Exception as e:
            utils.g_logger.info("数据填报-催报查询页显示异常")
            sheet.append(["数据填报-催报查询", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_催报查询.png", doc, "数据填报_催报查询")

        # 数据填报报送评价页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='报送评价']").click()
        try:
            time.sleep(10)
            self.driver.find_element(By.XPATH, "//div[@id='reportboxTree']//i[@class='el-icon-arrow-down']").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//li[starts-with(@id,'reportTree')][1]")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开报送评价页，且报表列表存在")
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开报送评价页，且站点列表存在")
            utils.g_logger.info("数据填报-报送评价页显示正常")
            sheet.append(["数据填报-报送评价", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_报送评价.png", doc, "数据填报_报送评价")
        except Exception as e:
            utils.g_logger.info("数据填报-报送评价页显示异常")
            sheet.append(["数据填报-报送评价", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_报送评价.png", doc, "数据填报_报送评价")

        # 数据填报填报任务一览页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='填报任务一览']").click()
        try:
            time.sleep(10)
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开填报任务一览页，且站点列表存在")
            utils.g_logger.info("数据填报-填报任务一览页显示正常")
            sheet.append(["数据填报-填报任务一览", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/填报任务一览.png", doc, "填报任务一览")
        except Exception as e:
            utils.g_logger.info("数据填报-填报任务一览页显示异常")
            sheet.append(["数据填报-填报任务一览", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/填报任务一览.png", doc, "填报任务一览")

        # 数据填报汇总统计页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='汇总统计']").click()
        try:
            time.sleep(10)
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开数据填报-汇总统计页，且站点列表存在")
            utils.g_logger.info("数据填报-汇总统计页显示正常")
            sheet.append(["数据填报-汇总统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_汇总统计.png", doc, "数据填报_汇总统计")
        except Exception as e:
            utils.g_logger.info("数据填报-汇总统计页异常")
            sheet.append(["数据填报-汇总统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_汇总统计.png", doc, "数据填报_汇总统计")

        # 数据填报任务审核页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='任务审核']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='search-container']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开数据填报-任务审核页，且指定元素存在")
            utils.g_logger.info("数据填报-任务审核页显示正常")
            sheet.append(["数据填报-任务审核", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_任务审核.png", doc, "数据填报_任务审核")
        except Exception as e:
            utils.g_logger.info("数据填报-任务审核页异常")
            sheet.append(["数据填报-任务审核", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据填报_任务审核.png", doc, "数据填报_任务审核")

        # 数据填报特色表分析页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='特色表分析']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='tab-information']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开数据填报-特色表分析页，且指定元素存在")
            utils.g_logger.info("数据填报-特色表分析页显示正常")
            sheet.append(["数据填报-特色表分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/特色表分析.png", doc, "特色表分析")
        except Exception as e:
            utils.g_logger.info("数据填报-特色表分析页异常")
            sheet.append(["数据填报-特色表分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/特色表分析.png", doc, "特色表分析")

        # 执行”关闭所有“页面操作
        self.driver.find_element(By.XPATH, "//i[@class='el-icon-circle-close']").click()
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='el-dropdown-menu el-popper']/li[text()='关闭所有']"))
            )
            element.click()
            utils.g_logger.info("成功关闭所有页面")
        except Exception as e:
            utils.g_logger.info(e)

        # 系统管理报表权限管理页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//div[starts-with(@class,'navi-item')][text()='系统管理 ']").click()
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='报表权限管理']").click()
        try:
            time.sleep(10)
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开报表权限管理页，且站点列表存在")
            utils.g_logger.info("系统管理-报表权限管理页显示正常")
            sheet.append(["系统管理-报表权限管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/报表权限管理.png", doc, "报表权限管理")
        except Exception as e:
            utils.g_logger.info("系统管理-报表权限管理页异常")
            sheet.append(["系统管理-报表权限管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/报表权限管理.png", doc, "报表权限管理")

        # 系统管理帮助管理页
        """self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='帮助管理']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='el-table__fixed-right']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开帮助管理页，且指定元素存在")
            utils.g_logger.info("系统管理-帮助管理页显示正常")
            sheet.append(["系统管理-帮助管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/帮助管理.png", doc, "帮助管理")
        except Exception as e:
            utils.g_logger.info("系统管理-帮助管理页异常")
            sheet.append(["系统管理-帮助管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/帮助管理.png", doc, "帮助管理")

        # 系统管理工作平台页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='工作平台']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='flexbox']//div[@class='title'][text()='省']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-工作平台页，且指定元素存在")
            utils.g_logger.info("系统管理-工作平台页显示正常")
            sheet.append(["系统管理-工作平台", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_工作平台.png", doc, "系统管理_工作平台")
        except Exception as e:
            utils.g_logger.info("系统管理-工作平台页异常")
            sheet.append(["系统管理-工作平台", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_工作平台.png", doc, "系统管理_工作平台")

        # 系统管理机构管理页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='机构管理']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@for='orglevel']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-机构管理页，且指定元素存在")
            utils.g_logger.info("系统管理-机构管理页显示正常")
            sheet.append(["系统管理-机构管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/机构管理.png", doc, "机构管理")
        except Exception as e:
            utils.g_logger.info("系统管理-机构管理页异常")
            sheet.append(["系统管理-机构管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/机构管理.png", doc, "机构管理")"""

        # 系统管理用户管理页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='用户管理']").click()
        try:
            time.sleep(13)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@for='username']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-用户管理页，且指定元素存在")
            utils.g_logger.info("系统管理-用户管理页显示正常")
            sheet.append(["系统管理-用户管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_用户管理.png", doc, "系统管理_用户管理")
        except Exception as e:
            utils.g_logger.info("系统管理-用户管理页异常")
            sheet.append(["系统管理-用户管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_用户管理.png", doc, "系统管理_用户管理")

        # 系统管理权限管理页
        """self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='权限管理']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='cell el-tooltip'][text()='系统管理员']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-权限管理页，且指定元素存在")
            utils.g_logger.info("系统管理-权限管理页显示正常")
            sheet.append(["系统管理-权限管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_权限管理.png", doc, "系统管理_权限管理")
        except Exception as e:
            utils.g_logger.info("系统管理-权限管理页异常")
            sheet.append(["系统管理-权限管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_权限管理.png", doc, "系统管理_权限管理")

        # 系统管理菜单管理页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='菜单管理']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@for='menuid']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-菜单管理页，且指定元素存在")
            utils.g_logger.info("系统管理-菜单管理页显示正常")
            sheet.append(["系统管理-菜单管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_菜单管理.png", doc, "系统管理_菜单管理")
        except Exception as e:
            utils.g_logger.info("系统管理-菜单管理页异常")
            sheet.append(["系统管理-菜单管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_菜单管理.png", doc, "系统管理_菜单管理")

        # 系统管理字典表管理页
        self.driver.find_element(By.XPATH, "//span[@slot='title' and text()='字典表管理']").click()
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='系统字典表管理 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='el-table__fixed-right']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-字典表管理页，且指定元素存在")
            utils.g_logger.info("系统管理-字典表管理页显示正常")
            sheet.append(["系统管理-字典表管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_字典表管理.png", doc, "系统管理_字典表管理")
        except Exception as e:
            utils.g_logger.info("系统管理-字典表管理页异常")
            sheet.append(["系统管理-字典表管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_字典表管理.png", doc, "系统管理_字典表管理")

        # 系统管理日志管理登录日志页
        self.driver.find_element(By.XPATH, "//span[@slot='title' and text()='日志管理']").click()
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='登录日志 ']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='登录时间']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-登录日志页，且指定元素存在")
            utils.g_logger.info("系统管理-登录日志页显示正常")
            sheet.append(["系统管理-登录日志", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_登录日志.png", doc, "系统管理_日志管理_登录日志")
        except Exception as e:
            utils.g_logger.info("系统管理-登录日志页异常")
            sheet.append(["系统管理-登录日志", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_登录日志.png", doc, "系统管理_日志管理_登录日志")

        # 系统管理日志管理操作日志页
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='操作日志 ']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='操作类型']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-操作日志页，且指定元素存在")
            utils.g_logger.info("系统管理-操作日志页显示正常")
            sheet.append(["系统管理-操作日志", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_操作日志.png", doc, "系统管理_日志管理_操作日志")
        except Exception as e:
            utils.g_logger.info("系统管理-操作日志页异常")
            sheet.append(["系统管理-操作日志", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_操作日志.png", doc, "系统管理_日志管理_操作日志")

        # 系统管理日志管理上报日志页
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='上报日志 ']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='上报国家状态']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-上报日志页，且指定元素存在")
            utils.g_logger.info("系统管理-上报日志页显示正常")
            sheet.append(["系统管理-上报日志", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_上报日志.png", doc, "系统管理_日志管理_上报日志")
        except Exception as e:
            utils.g_logger.info("系统管理-上报日志页异常")
            sheet.append(["系统管理-上报日志", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_上报日志.png", doc, "系统管理_日志管理_上报日志")

        # 系统管理日志管理同步日志页
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='同步日志 ']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='同步时间']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-同步日志页，且指定元素存在")
            utils.g_logger.info("系统管理-同步日志页显示正常")
            sheet.append(["系统管理-同步日志", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_同步日志.png", doc, "系统管理_日志管理_同步日志")
        except Exception as e:
            utils.g_logger.info("系统管理-同步日志页异常")
            sheet.append(["系统管理-同步日志", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_日志管理_同步日志.png", doc, "系统管理_日志管理_同步日志")"""

        # 系统管理填报任务设置页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='填报任务设置']").click()
        try:
            time.sleep(13)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//button/span[text()=' 新增任务 ']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-填报任务设置页，且指定元素存在")
            utils.g_logger.info("系统管理-填报任务设置页显示正常")
            sheet.append(["系统管理-填报任务设置", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_填报任务设置.png", doc,
                                  "系统管理_填报任务设置")
        except Exception as e:
            utils.g_logger.info("系统管理-填报任务设置异常")
            sheet.append(["系统管理-填报任务设置", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_填报任务设置.png", doc,
                                  "系统管理_填报任务设置")

        # 系统管理专业分析配置页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='专业分析配置']").click()
        try:
            time.sleep(13)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='分析类型']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-专业分析配置页，且指定元素存在")

            utils.g_logger.info("系统管理-专业分析配置页显示正常")
            sheet.append(["系统管理-专业分析配置", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_专业分析配置.png", doc,
                                  "系统管理_专业分析配置")
        except Exception as e:
            utils.g_logger.info("系统管理-专业分析配置页异常")
            sheet.append(["系统管理-专业分析配置", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_填报任务设置.png", doc,
                                  "系统管理_专业分析配置")

        # 系统管理定时任务管理页
        """self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='定时任务']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//thead[@class='has-gutter']//div[text()='任务名称']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-定时任务管理页，且指定元素存在")
            utils.g_logger.info("系统管理-定时任务管理页显示正常")
            sheet.append(["系统管理-定时任务管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_定时任务管理.png", doc,
                                  "系统管理_定时任务管理")
        except Exception as e:
            utils.g_logger.info("系统管理-定时任务管理页异常")
            sheet.append(["系统管理-定时任务管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_定时任务管理.png", doc,
                                  "系统管理_定时任务管理")

        # 系统管理定制报告页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='定制报告']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='周期']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-定制报告页，且指定元素存在")
            utils.g_logger.info("系统管理-定制报告页显示正常")
            sheet.append(["系统管理-定制报告", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_定制报告.png", doc,
                                  "系统管理_定制报告")
        except Exception as e:
            utils.g_logger.info("系统管理-定制报告页异常")
            sheet.append(["系统管理-定制报告", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_填报任务设置.png", doc,
                                  "系统管理_定制报告")

        # 系统管理预警阈值设置页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='预警阈值设置']").click()
        try:
            time.sleep(13)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='分析指标']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-预警阈值设置页，且指定元素存在")
            utils.g_logger.info("系统管理-预警阈值设置页显示正常")
            sheet.append(["系统管理-预警阈值设置", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_预警阈值设置.png", doc,
                                  "系统管理_预警阈值设置")
        except Exception as e:
            utils.g_logger.info("系统管理-预警阈值设置页异常")
            sheet.append(["系统管理-预警阈值设置", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_预警阈值设置.png", doc,
                                  "系统管理_预警阈值设置")

        # 系统管理模型管理页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='模型管理']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//th[starts-with(@class,'el-table_')]/div[text()='模型名称'][1]")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-模型管理页，且指定元素存在")
            utils.g_logger.info("系统管理-模型管理页显示正常")
            sheet.append(["系统管理-模型管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_模型管理.png", doc,
                                  "系统管理_模型管理")
        except Exception as e:
            utils.g_logger.info("系统管理-模型管理页异常")
            sheet.append(["系统管理-模型管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_模型管理.png", doc,
                                  "系统管理_模型管理")

        # 系统管理县级用户关系绑定页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='县级用户关系绑定']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//th[starts-with(@class,'el-table_')]/div[text()='县级系统用户登录名']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-县级用户关系绑定页，且指定元素存在")
            utils.g_logger.info("系统管理-县级用户关系绑定页显示正常")
            sheet.append(["系统管理-县级用户关系绑定", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_县级用户关系绑定.png", doc,
                                  "系统管理_县级用户关系绑定")
        except Exception as e:
            utils.g_logger.info("系统管理-县级用户关系绑定页异常")
            sheet.append(["系统管理-县级用户关系绑定", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_县级用户关系绑定.png", doc,
                                  "系统管理_县级用户关系绑定")

        # 系统管理系统更新日志页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'el-menu-item first-menu')]//span[text()='系统更新日志']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//th[starts-with(@class,'el-table_')]/div[text()='版本号']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开系统管理-系统更新日志页，且指定元素存在")
            utils.g_logger.info("系统管理-系统更新日志页显示正常")
            sheet.append(["系统管理-系统更新日志", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_系统更新日志.png", doc,
                                  "系统管理_系统更新日志")
        except Exception as e:
            utils.g_logger.info("系统管理-系统更新日志页异常")
            sheet.append(["系统管理-系统更新日志", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/系统管理_系统更新日志.png", doc,
                                  "系统管理_系统更新日志")"""

        # 关闭所有页面
        self.driver.find_element(By.XPATH, "//i[@class='el-icon-circle-close']").click()
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='el-dropdown-menu el-popper']/li[text()='关闭所有']"))
            )
            element.click()
            utils.g_logger.info("成功关闭所有页面")
        except Exception as e:
            print(e)

        # 物联网工作平台页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//div[starts-with(@class,'navi-item')][text()='物联网 ']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='dashboard_widget_box']/div[@class='widget_title widget_title_heading'][text()=' 物联网统计分析 ']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开物联网-工作平台页，且指定元素存在")
            utils.g_logger.info("物联网-工作平台页显示正常")
            sheet.append(["物联网-工作平台", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_工作平台.png", doc, "物联网_工作平台")
        except Exception as e:
            utils.g_logger.info("物联网-工作平台页异常")
            sheet.append(["物联网-工作平台", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_工作平台.png", doc, "物联网_工作平台")

        # 物联网-监测点分布页面
        self.driver.find_element(By.XPATH, "//li[@class='el-menu-item first-menu']//span[text()='监测点分布']").click()
        try:
            time.sleep(15)
            tree_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='equipmentbox']/div[1]/div[2]/div[1]")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开监测点分布页面，且指定元素存在")
            utils.g_logger.info("物联网-监测点分布显示正常")
            sheet.append(["物联网-监测点分布", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_监测点分布.png", doc, "物联网_监测点分布")
        except Exception as e:
            utils.g_logger.info("物联网-监测点分布页异常")
            sheet.append(["物联网-监测点分布", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_监测点分布.png", doc, "物联网_监测点分布")

        # 物联网-设备分布页面
        self.driver.find_element(By.XPATH, "//li[@class='el-menu-item first-menu']//span[text()='设备分布']").click()
        self.driver.find_element(By.XPATH, "//*[@id='tab-detail']").click()
        try:
            time.sleep(15)
            tree_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='pane-detail']/div/div[2]/div/div")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开设备分布页面，且指定元素存在")
            utils.g_logger.info("物联网-设备分布显示正常")
            sheet.append(["物联网-设备分布", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_设备分布.png", doc, "物联网_设备分布")
        except Exception as e:
            utils.g_logger.info("物联网-设备分布页异常")
            sheet.append(["物联网-设备分布", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_设备分布.png", doc, "物联网_设备分布")

        # 物联网-环境气象-趋势分析
        self.driver.find_element(By.XPATH, "//div[@class='el-submenu__title']//span[text()='环境气象']").click()
        self.driver.find_element(By.XPATH,
                                 "//li[@class='el-submenu is-opened']//li[text()='趋势分析 ']").click()
        try:
            time.sleep(14)
            tree_element = WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, "//span[@name='qixiangqushifenxi']")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开环境气象-趋势分析页面，且指定元素存在")
            utils.g_logger.info("环境气象-趋势分析显示正常")
            sheet.append(["环境气象-趋势分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_趋势分析.png", doc,
                                  "物联网_环境气象_趋势分析")
        except Exception as e:
            utils.g_logger.info("环境气象-趋势分析页异常")
            sheet.append(["环境气象-趋势分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_趋势分析.png", doc,
                                  "物联网_环境气象_趋势分析")

        # 物联网-环境气象-实时数据列表
        self.driver.find_element(By.XPATH, "//li[text()='实时数据列表 ']").click()
        try:
            time.sleep(15)
            tree_element = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='el-table__body-wrapper is-scrolling-none']/table[@class='el-table__body']")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开环境气象-实时数据列表页面，且指定元素存在")
            utils.g_logger.info("环境气象-实时数据列表显示正常")
            sheet.append(["环境气象-实时数据列表", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_实时数据列表.png", doc,
                                  "物联网_环境气象_实时数据列表")
        except Exception as e:
            utils.g_logger.info("环境气象-实时数据列表显示正常")
            sheet.append(["环境气象-实时数据列表", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_实时数据列表.png", doc,
                                  "物联网_环境气象_实时数据列表")

        # 物联网-环境气象-实时数据统计
        self.driver.find_element(By.XPATH, "//li[text()='实时数据统计 ']").click()
        try:
            time.sleep(15)
            tree_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='devicenew']/div[1]/div[4]/button/span[text()='可选指标']")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开环境气象-实时数据统计页面，且指定元素存在")
            utils.g_logger.info("环境气象-实时数据统计显示正常")
            sheet.append(["环境气象-实时数据统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_实时数据统计.png", doc,
                                  "物联网_环境气象_实时数据统计")
        except Exception as e:
            utils.g_logger.info("环境气象-实时数据统计页异常")
            sheet.append(["环境气象-实时数据统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_实时数据统计.png", doc,
                                  "物联网_环境气象_实时数据统计")

        # 物联网-环境气象-逐日数据统计
        self.driver.find_element(By.XPATH, "//li[text()='逐日数据统计 ']").click()
        try:
            time.sleep(15)
            tree_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='elmain']/div[2]/div/div[1]/div[4]/button//span[text()='可选指标']")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开环境气象-逐日数据统计页面，且指定元素存在")
            utils.g_logger.info("环境气象-逐日数据统计显示正常")
            sheet.append(["环境气象-逐日数据统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_逐日数据统计.png", doc,
                                  "物联网_环境气象_逐日数据统计")
        except Exception as e:
            utils.g_logger.info("环境气象-逐日数据统计页异常")
            sheet.append(["环境气象-逐日数据统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_逐日数据统计.png", doc,
                                  "物联网_环境气象_逐日数据统计")

        # 物联网-环境气象-逐日数据列表
        self.driver.find_element(By.XPATH, "//li[text()='逐日数据列表 ']").click()
        try:
            time.sleep(14)
            wendu_element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//th[contains(@class,'     is-leaf el-table__cell')]//div[@class='cell'][text()='平均气温（°C）']")))
            unittest.TestCase.assertTrue(wendu_element is not None,
                                         "成功打开环境气象-逐日数据列表页，且站点列表存在")
            utils.g_logger.info("环境气象-逐日数据列表显示正常")
            sheet.append(["环境气象-逐日数据列表", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_逐日数据列表.png", doc,
                                  "物联网_环境气象_逐日数据列表")
        except Exception as e:
            utils.g_logger.info("环境气象-逐日数据列表页异常")
            sheet.append(["环境气象-逐日数据列表", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_环境气象_逐日数据列表.png", doc,
                                  "物联网_环境气象_逐日数据列表")

        # 性诱监测-性诱数据分析
        self.driver.find_element(By.XPATH, "//div[@class='el-submenu__title']//span[text()='性诱监测']").click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//li[text()='性诱数据分析 ']").click()
        try:
            time.sleep(13)
            tree_element = WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='ol-zoom ol-unselectable ol-control']")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开性诱监测-性诱数据分析页面，且指定元素存在")
            utils.g_logger.info("性诱监测-性诱数据分析显示正常")
            sheet.append(["性诱监测-性诱数据分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_性诱数据分析.png", doc,
                                  "物联网_性诱监测_性诱数据分析")
        except Exception as e:
            utils.g_logger.info("性诱监测-性诱数据分析页异常")
            sheet.append(["性诱监测-性诱数据分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_性诱数据分析.png", doc,
                                  "物联网_性诱监测_性诱数据分析")

        # 性诱监测-数据统计列表
        self.driver.find_element(By.XPATH,
                                 "//li[@class='el-submenu is-active is-opened']//li[text()='数据统计列表 ']").click()
        try:
            time.sleep(13)
            tree_element = WebDriverWait(self.driver, 15).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='el-col_label']")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开性诱监测-数据统计列表页面，且指定元素存在")
            utils.g_logger.info("性诱监测-数据统计列表显示正常")
            sheet.append(["性诱监测-数据统计列表", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_数据统计列表.png", doc,
                                  "物联网_性诱监测_数据统计列表")
        except Exception as e:
            utils.g_logger.info("性诱监测-数据统计列表页异常")
            sheet.append(["性诱监测-数据统计列表", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_数据统计列表.png", doc,
                                  "物联网_性诱监测_数据统计列表")

        # 性诱监测-性诱数据统计
        self.driver.find_element(By.XPATH, "//li[text()='性诱数据统计 ']").click()
        try:
            time.sleep(14)
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开报表权限管理页，且站点列表存在")
            utils.g_logger.info("性诱监测-性诱数据统计显示正常")
            sheet.append(["性诱监测-性诱数据统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_性诱数据统计.png", doc,
                                  "物联网_性诱监测_性诱数据统计")
        except Exception as e:
            utils.g_logger.info("性诱监测-性诱数据统计页异常")
            sheet.append(["性诱监测-性诱数据统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_性诱数据统计.png", doc,
                                  "物联网_性诱监测_性诱数据统计")

        # 性诱监测-趋势分析
        self.driver.find_element(By.XPATH,
                                 "//div[@class='el-submenu__title']//span[text()='性诱监测']/parent::div/parent::li//li[text()='趋势分析 ']").click()
        try:
            time.sleep(13)
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功性诱监测-趋势分析页，且站点列表存在")
            tree_element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//label[@class='el-form-item__label'][text()='作物类型']/following-sibling::div")))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开性诱监测-趋势分析页面，且作物类型选项存在")
            utils.g_logger.info("性诱监测-趋势分析显示正常")
            sheet.append(["性诱监测-趋势分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_趋势分析.png", doc,
                                  "物联网_性诱监测_趋势分析")
        except Exception as e:
            utils.g_logger.info("性诱监测-趋势分析页异常")
            sheet.append(["性诱监测-趋势分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_性诱监测_趋势分析.png", doc,
                                  "物联网_性诱监测_趋势分析")

        # 灯诱监测-灯诱数据分析页
        self.driver.find_element(By.XPATH, "//span[@slot='title' and text()='灯诱监测']").click()
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='灯诱数据分析 ']").click()
        try:
            time.sleep(14)
            element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//div[@id='timeboxx']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-灯诱监测-灯诱数据分析页，且指定元素存在")
            utils.g_logger.info("物联网-灯诱监测-灯诱数据分析页显示正常")
            sheet.append(["物联网-灯诱监测-灯诱数据分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_灯诱数据分析.png", doc,
                                  "物联网_灯诱监测_灯诱数据分析")
        except Exception as e:
            utils.g_logger.info("物联网-灯诱监测-灯诱数据分析页异常")
            sheet.append(["物联网-灯诱监测-灯诱数据分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_灯诱数据分析.png", doc,
                                  "物联网_灯诱监测_灯诱数据分析")

        # 灯诱监测-数据统计列表
        self.driver.find_element(By.XPATH,
                                 "//span[@slot='title' and text()='灯诱监测']/parent::div/parent::li//li[text()='数据统计列表 ']").click()
        try:
            time.sleep(14)
            element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//thead[@class='has-gutter']//div[text()='日累计虫量']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-灯诱监测-数据统计列表页，且指定元素存在")
            utils.g_logger.info("物联网-灯诱监测-数据统计列表页显示正常")
            sheet.append(["物联网-灯诱监测-数据统计列表", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_数据统计列表.png", doc,
                                  "物联网_灯诱监测_数据统计列表")
        except Exception as e:
            utils.g_logger.info("物联网-灯诱监测-数据统计列表页异常")
            sheet.append(["物联网-灯诱监测-数据统计列表", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_数据统计列表.png", doc,
                                  "物联网_灯诱监测_数据统计列表")

        # 灯诱监测-灯诱图片展示
        self.driver.find_element(By.XPATH,
                                 "//li[@role='menuitem'][text()='灯诱图片展示 ']").click()
        try:
            time.sleep(16)
            calendar_element = WebDriverWait(self.driver, 15).until(EC.visibility_of_element_located(
                (By.XPATH, "//table[@class='el-calendar-table']")))
            unittest.TestCase.assertTrue(calendar_element is not None,
                                         "成功打开物联网-灯诱监测-灯诱图片展示页，且日历元素存在")
            image_element = WebDriverWait(self.driver, 15).until(EC.visibility_of_element_located(
                (By.XPATH, "//div[@id='map']/div[@class='time']")))
            unittest.TestCase.assertTrue(image_element is not None,
                                         "成功打开物联网-灯诱监测-灯诱图片展示页，且图片元素存在")
            utils.g_logger.info("物联网-灯诱监测-灯诱图片展示页显示正常")
            sheet.append(["物联网-灯诱监测-灯诱图片展示", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_灯诱图片展示.png", doc,
                                  "物联网_灯诱监测_灯诱图片展示")
        except Exception as e:
            utils.g_logger.info("物联网-灯诱监测-灯诱图片展示页异常")
            sheet.append(["物联网-灯诱监测-灯诱图片展示", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_灯诱图片展示.png", doc,
                                  "物联网_灯诱监测_灯诱图片展示")

        # 灯诱监测-灯诱识别结果统计
        self.driver.find_element(By.XPATH,
                                 "//li[@role='menuitem'][text()='灯诱识别结果统计 ']").click()
        try:
            time.sleep(14)
            element = WebDriverWait(self.driver, 15).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='el-col_label']/parent::div/following-sibling::div/div[@class='el-checkbox-group']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-灯诱监测-灯诱识别结果统计页，且虫害类型选项存在")
            utils.g_logger.info("物联网-灯诱监测-灯诱识别结果统计页显示正常")
            sheet.append(["物联网-灯诱监测-灯诱识别结果统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_灯诱识别结果统计.png", doc,
                                  "物联网_灯诱监测_灯诱识别结果统计")
        except Exception as e:
            utils.g_logger.info("物联网-灯诱监测-灯诱识别结果统计页异常")
            sheet.append(["物联网-灯诱监测-灯诱识别结果统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_灯诱识别结果统计.png", doc,
                                  "物联网_灯诱监测_灯诱识别结果统计")

        # 灯诱监测-趋势分析
        time.sleep(5)
        self.driver.find_element(By.XPATH,
                                 "//li[@role='menuitem'][text()='灯诱识别结果统计 ']/parent::div/following-sibling::div").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 15).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='el-col_label']/parent::div/following-sibling::div/div[@role='radiogroup']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-灯诱监测-趋势分析页，且虫害类型选项存在")
            utils.g_logger.info("物联网-灯诱监测-趋势分析页显示正常")
            sheet.append(["物联网-灯诱监测-趋势分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_趋势分析.png", doc,
                                  "物联网_灯诱监测_趋势分析")
        except Exception as e:
            utils.g_logger.info("物联网-灯诱监测-趋势分析页异常")
            sheet.append(["物联网-灯诱监测-趋势分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_灯诱监测_趋势分析.png", doc,
                                  "物联网_灯诱监测_趋势分析")

        # 病害监测-马铃薯晚疫病页
        time.sleep(15)
        self.driver.find_element(By.XPATH, "//span[@slot='title' and text()='病害监测']").click()
        time.sleep(4)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='马铃薯晚疫病 ']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 15).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='legendbox']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-病害监测-马铃薯晚疫病页，且病害分布图图例存在")
            utils.g_logger.info("物联网-病害监测-马铃薯晚疫病页显示正常")
            sheet.append(["物联网-病害监测-马铃薯晚疫病", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_病害监测_马铃薯晚疫病.png", doc,
                                  "物联网_病害监测_马铃薯晚疫病")
        except Exception as e:
            utils.g_logger.info("物联网-病害监测-马铃薯晚疫病页异常")
            sheet.append(["物联网-病害监测-马铃薯晚疫病", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_病害监测_马铃薯晚疫病.png", doc,
                                  "物联网_病害监测_马铃薯晚疫病")

        # 病害监测-小麦赤霉病页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='小麦赤霉病 ']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 15).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='legendbox']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-病害监测-小麦赤霉病页，且病害分布图图例存在")
            utils.g_logger.info("物联网-病害监测-小麦赤霉病页显示正常")
            sheet.append(["物联网-病害监测-小麦赤霉病", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_病害监测_小麦赤霉病.png", doc,
                                  "物联网_病害监测_小麦赤霉病")
        except Exception as e:
            utils.g_logger.info("物联网-病害监测-小麦赤霉病页异常")
            sheet.append(["物联网-病害监测-小麦赤霉病", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_病害监测_小麦赤霉病.png", doc,
                                  "物联网_病害监测_小麦赤霉病")

        # 病害监测-孢子监测
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='孢子监测 ']").click()
        try:
            time.sleep(13)
            element = WebDriverWait(self.driver, 15).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@id='map']/div[@class='time']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-病害监测-孢子监测页，且孢子图片存在")
            utils.g_logger.info("物联网-病害监测-孢子监测页显示正常")
            sheet.append(["物联网-病害监测-孢子监测", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_病害监测_孢子监测.png", doc,
                                  "物联网_病害监测_孢子监测")
        except Exception as e:
            utils.g_logger.info("物联网-病害监测-孢子监测页异常")
            sheet.append(["物联网-病害监测-孢子监测", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_病害监测_孢子监测.png", doc,
                                  "物联网_病害监测_孢子监测")

        # 关闭其他页面
        self.driver.find_element(By.XPATH, "//span[text()='工作平台']").click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//i[@class='el-icon-circle-close']").click()
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='el-dropdown-menu el-popper']/li[text()='关闭其他']"))
            )
            element.click()
            utils.g_logger.info("成功关闭其他页面")
        except Exception as e:
            print(e)

        # 虫量对比分析页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='虫量对比分析']").click()
        try:
            time.sleep(14)
            element = WebDriverWait(self.driver, 15).until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//label[text()='虫害类型']/following-sibling::div//div[@role='radiogroup']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-虫量对比分析页，且虫害类型选项存在")
            utils.g_logger.info("物联网-虫量对比分析页显示正常")
            sheet.append(["物联网-虫量对比分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_虫量对比分析.png", doc,
                                  "物联网_虫量对比分析")
        except Exception as e:
            utils.g_logger.info("物联网-虫量对比分析页异常")
            sheet.append(["物联网-虫量对比分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_虫量对比分析.png", doc,
                                  "物联网_虫量对比分析")

        # 物联网-物联网管理-设备管理页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//span[@slot='title' and text()='物联网管理']").click()
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='设备管理 ']").click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='设备管理 ']").click()
        try:
            time.sleep(16)
            element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//label[@for='equipmenttypename']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-物联网管理-设备管理页，且指定元素存在")
            utils.g_logger.info("物联网-物联网管理-设备管理页显示正常")
            sheet.append(["物联网-物联网管理-设备管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_物联网管理_设备管理.png", doc,
                                  "物联网_物联网管理_设备管理")
        except Exception as e:
            utils.g_logger.info("物联网-物联网管理-设备管理页异常")
            sheet.append(["物联网-物联网管理-设备管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_物联网管理_设备管理.png", doc,
                                  "物联网_物联网管理_设备管理")

        # 物联网-物联网管理-监测点管理页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='监测点管理 ']").click()
        try:
            time.sleep(16)
            element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//label[@for='equipmentname'][text()='监测点名称']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开物联网-物联网管理-监测点管理页，且指定元素存在")
            utils.g_logger.info("物联网-物联网管理-监测点管理页显示正常")
            sheet.append(["物联网-物联网管理-监测点管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_物联网管理_监测点管理.png", doc,
                                  "物联网_物联网管理_监测点管理")
        except Exception as e:
            utils.g_logger.info("物联网-物联网管理-监测点管理页异常")
            sheet.append(["物联网-物联网管理-监测点管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_物联网管理_监测点管理.png", doc,
                                  "物联网_物联网管理_监测点管理")

        # 物联网-视频监控-视频监控分布
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//span[@slot='title' and text()='视频监控']").click()
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='视频监控分布 ']").click()
        try:
            time.sleep(13)
            self.driver.find_element(By.XPATH, "//div[@id='orgboxzTree']//i[@class='el-icon-arrow-down']").click()
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//ul[starts-with(@id,'orgTree') and @class='ztree']//a[@title='广西壮族自治区']")))
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开物联网-视频监控分布页，且站点列表存在")
            utils.g_logger.info("物联网-视频监控-视频监控分布页显示正常")
            sheet.append(["物联网-视频监控-视频监控分布", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_视频监控_视频监控分布.png", doc,
                                  "物联网_视频监控_视频监控分布")
        except Exception as e:
            utils.g_logger.info("物联网-视频监控-视频监控分布页异常")
            sheet.append(["物联网-视频监控-视频监控分布", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_视频监控_视频监控分布.png", doc,
                                  "物联网_视频监控_视频监控分布")

        # 物联网-视频监控-视频图片展示
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='视频图片展示 ']").click()
        try:
            time.sleep(15)
            calendar_element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//table[@class='el-calendar-table']")))
            unittest.TestCase.assertTrue(calendar_element is not None,
                                         "成功打开物联网-视频监控-视频图片展示页，且日历元素存在")
            image_element = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//div[@id='map']/div[@class='time']")))
            unittest.TestCase.assertTrue(image_element is not None,
                                         "成功打开物联网-视频监控-视频图片展示页，且图片元素存在")
            utils.g_logger.info("物联网-视频监控-视频图片展示页显示正常")
            sheet.append(["物联网-视频监控-视频图片展示", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_视频监控_视频图片展示.png", doc,
                                  "物联网_视频监控_视频图片展示")
        except Exception as e:
            utils.g_logger.info("物联网-视频监控-视频图片展示页异常")
            sheet.append(["物联网-视频监控-视频图片展示", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/物联网_视频监控_视频图片展示.png", doc,
                                  "物联网_视频监控_视频图片展示")

        # 关闭所有页面
        self.driver.find_element(By.XPATH, "//i[@class='el-icon-circle-close']").click()
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='el-dropdown-menu el-popper']/li[text()='关闭所有']"))
            )
            element.click()
            utils.g_logger.info("成功关闭所有页面")
        except Exception as e:
            print(e)

        # 模型预警页面
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//div[starts-with(@class,'navi-item')][text()='模型预警 ']").click()
        try:
            time.sleep(15)
            self.driver.find_element(By.XPATH, "//label[text()='模型名称']/following-sibling::div//input").click()
            element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//div[@x-placement]//ul[@class='el-scrollbar__view el-select-dropdown__list']/li[1]")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开模型预警页，且指定元素存在")
            utils.g_logger.info("模型预警页显示正常")
            sheet.append(["模型预警", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/模型预警.png", doc, "模型预警")
        except Exception as e:
            utils.g_logger.info("模型预警页异常")
            sheet.append(["模型预警", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/模型预警.png", doc, "模型预警")

        # 数据分析-综合分析页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//div[starts-with(@class,'navi-item')][text()='数据分析 ']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@role='radiogroup']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开数据分析-综合分析页，且指定元素存在")
            select_date_element = self.driver.find_element(By.XPATH, '//input[@placeholder="请选择日期"]')
            select_date_element.clear()
            select_date_element.send_keys('2024-10-08')
            time.sleep(3)
            # 模拟按下键盘回车键
            select_date_element.send_keys(Keys.RETURN)

            time.sleep(15)
            chart_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//div[text()=' 当前发生面积与常年对比 ']/parent::div//canvas")))
            unittest.TestCase.assertTrue(chart_element is not None, "成功打开数据分析-综合分析页，且指定元素存在")
            utils.g_logger.info("数据分析-综合分析页显示正常")
            sheet.append(["数据分析-综合分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_综合分析.png", doc, "数据分析_综合分析")
        except Exception as e:
            utils.g_logger.info("数据分析-综合分析页异常")
            sheet.append(["数据分析-综合分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_综合分析.png", doc, "数据分析_综合分析")

        # 数据分析-专题分析页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='专题分析']").click()
        try:
            time.sleep(15)
            elements = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located(
                (By.XPATH, "//div[@class='el-image']")))
            unittest.TestCase.assertTrue(elements is not None,
                                         "成功打开数据分析-专题分析页，且指定元素存在")
            self.driver.find_element(By.XPATH, "//div[@class='row-title'][text()=' 稻纵卷叶螟 ']").click()
            time.sleep(5)
            indicator_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//div[@class='section_left section_left_light']/ul/li[1]")))
            unittest.TestCase.assertTrue(indicator_element is not None,
                                         "成功打开数据分析-专题分析页，且指定元素存在")
            utils.g_logger.info("数据分析-专题分析页显示正常")
            sheet.append(["数据分析-专题分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_专题分析.png", doc, "数据分析_专题分析")
            self.driver.find_element(By.XPATH, "//div[@class='el-form-item__content']//span[text()='关闭']").click()
        except Exception as e:
            utils.g_logger.info("数据分析-专题分析页异常")
            sheet.append(["数据分析-专题分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_专题分析.png", doc, "数据分析_专题分析")

        # 数据分析-GIS分析页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='GIS分析']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//label[text()='分析指标']")))
            unittest.TestCase.assertTrue(element is not None,
                                         "成功打开数据分析-GIS分析页，且指定元素存在")
            map_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//div[@class='map']")))
            unittest.TestCase.assertTrue(map_element is not None,
                                         "成功打开数据分析-GIS分析页，且指定元素存在")
            utils.g_logger.info("数据分析-GIS分析页显示正常")
            sheet.append(["数据分析-GIS分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_GIS分析.png", doc, "数据分析_GIS分析")
        except Exception as e:
            utils.g_logger.info("数据分析-GIS分析页异常")
            sheet.append(["数据分析-GIS分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_GIS分析.png", doc, "数据分析_GIS分析")

        # 数据分析-自定义分析页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='自定义分析']").click()
        try:
            time.sleep(8)
            self.driver.find_element(By.XPATH, "//div[@class='report-selector']").click()
            elements = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//img[contains(@src,'/upload/specialtype/')]")))
            unittest.TestCase.assertTrue(elements is not None, "成功打开数据分析-自定义分析页，且指定元素存在")
            utils.g_logger.info("数据分析-自定义分析页显示正常")
            sheet.append(["数据分析-自定义分析", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_自定义分析.png", doc, "数据分析_自定义分析")
        except Exception as e:
            utils.g_logger.info("数据分析-自定义分析页异常")
            sheet.append(["数据分析-自定义分析", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_自定义分析.png", doc, "数据分析_自定义分析")

        # 数据分析-数据报告页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='数据报告']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//label[text()='周期']/following-sibling::div/div[@role='radiogroup']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开数据分析-数据报告页，且指定元素存在")
            utils.g_logger.info("数据分析-数据报告页显示正常")
            sheet.append(["数据分析-数据报告", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_数据报告.png", doc, "数据分析_数据报告")
        except Exception as e:
            utils.g_logger.info("数据分析-数据报告页异常")
            sheet.append(["数据分析-数据报告", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/数据分析_数据报告.png", doc, "数据分析_数据报告")

        # 关闭所有页面
        self.driver.find_element(By.XPATH, "//i[@class='el-icon-circle-close']").click()
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='el-dropdown-menu el-popper']/li[text()='关闭所有']"))
            )
            element.click()
            utils.g_logger.info("成功关闭所有页面")
        except Exception as e:
            utils.g_logger.info(e)

        # 知识库-工作平台页
        self.driver.find_element(By.XPATH, "//div[starts-with(@class,'navi-item')][text()='知识库 ']").click()
        try:
            time.sleep(12)
            elements = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, "//div[@class='title'][text()='病虫害知识库']/parent::div/following-sibling::div//img")))
            unittest.TestCase.assertTrue(elements is not None, "成功打开知识库-工作平台页，且指定元素存在")
            utils.g_logger.info("知识库-工作平台页显示正常")
            sheet.append(["知识库-工作平台", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_工作平台.png", doc, "知识库_工作平台")
        except Exception as e:
            utils.g_logger.info("知识库-工作平台页异常")
            sheet.append(["知识库-工作平台", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_工作平台.png", doc, "知识库_工作平台")

        # 知识库-病虫害知识库-知识浏览页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='病虫害知识库']").click()
        self.driver.find_element(By.XPATH,
                                 "//span[text()='病虫害知识库']/parent::div/following-sibling::ul//li[text()='知识浏览 ']").click()
        try:
            time.sleep(12)
            elements = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//div[@class='el-image img']/img")))
            unittest.TestCase.assertTrue(elements is not None, "成功打开知识库-病虫害知识库-知识浏览页，且指定元素存在")
            utils.g_logger.info("知识库-病虫害知识库-知识浏览页显示正常")
            sheet.append(["知识库-病虫害知识库-知识浏览", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_病虫害知识库_知识浏览.png", doc,
                                  "知识库_病虫害知识库_知识浏览")
        except Exception as e:
            utils.g_logger.info("知识库-病虫害知识库-知识浏览页异常")
            sheet.append(["知识库-病虫害知识库-知识浏览", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_病虫害知识库_知识浏览.png", doc,
                                  "知识库_病虫害知识库_知识浏览")

        # 知识库-病虫害知识库-知识维护页
        time.sleep(3)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='病虫害知识库']/parent::div/following-sibling::ul//li[text()='知识维护 ']").click()
        try:
            time.sleep(9)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='修改时间']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开知识库-病虫害知识库-知识维护页，且指定元素存在")
            utils.g_logger.info("知识库-病虫害知识库-知识维护页显示正常")
            sheet.append(["知识库-病虫害知识库-知识维护", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_病虫害知识库_知识维护.png", doc,
                                  "知识库_病虫害知识库_知识维护")
        except Exception as e:
            utils.g_logger.info("知识库-病虫害知识库-知识维护页异常")
            sheet.append(["知识库-病虫害知识库-知识维护", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_病虫害知识库_知识维护.png", doc,
                                  "知识库_病虫害知识库_知识维护")

        # 植保知识库-知识浏览页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='植保知识库']").click()
        self.driver.find_element(By.XPATH,
                                 "//span[text()='植保知识库']/parent::div/following-sibling::ul//li[text()='知识浏览 ']").click()
        try:
            time.sleep(8)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@role='radiogroup']//span[text()='植物检疫']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开知识库-植保知识库-知识浏览页，且指定元素存在")
            utils.g_logger.info("知识库-植保知识库-知识浏览页显示正常")
            sheet.append(["知识库-植保知识库-知识浏览", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_植保知识库_知识浏览.png", doc,
                                  "知识库_植保知识库_知识浏览")
        except Exception as e:
            utils.g_logger.info("知识库-植保知识库-知识浏览页异常")
            sheet.append(["知识库-植保知识库-知识浏览", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_植保知识库_知识浏览.png", doc,
                                  "知识库_植保知识库_知识浏览")

        # 植保知识库-知识审核页
        time.sleep(3)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='植保知识库']/parent::div/following-sibling::ul//li[text()='知识审核 ']").click()
        try:
            time.sleep(8)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='发布时间']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开知识库-植保知识库-知识审核页，且指定元素存在")
            utils.g_logger.info("知识库-植保知识库-知识审核页显示正常")
            sheet.append(["知识库-植保知识库-知识审核", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_植保知识库_知识审核.png", doc,
                                  "知识库_植保知识库_知识审核")
        except Exception as e:
            utils.g_logger.info("知识库-植保知识库-知识审核页异常")
            sheet.append(["知识库-植保知识库-知识审核", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_植保知识库_知识审核.png", doc,
                                  "知识库_植保知识库_知识审核")

        # 植知识库-知识上传页
        time.sleep(3)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='植保知识库']/parent::div/following-sibling::ul//li[text()='知识上传 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='发布时间']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开知识库-植保知识库-知识上传页，且指定元素存在")
            utils.g_logger.info("知识库-植保知识库-知识上传页显示正常")
            sheet.append(["知识库-植保知识库-知识上传", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_植保知识库_知识上传.png", doc,
                                  "知识库_植保知识库_知识上传")
        except Exception as e:
            utils.g_logger.info("知识库-植保知识库-知识上传页异常")
            sheet.append(["知识库-植保知识库-知识上传", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_植保知识库_知识上传.png", doc,
                                  "知识库_植保知识库_知识上传")

        # 资料库页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='资料库']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='目录']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开知识库-资料库页，且指定元素存在")
            utils.g_logger.info("知识库-资料库页显示正常")
            sheet.append(["知识库-资料库", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_资料库.png", doc,
                                  "知识库_资料库")
        except Exception as e:
            utils.g_logger.info("知识库-资料库页异常")
            sheet.append(["知识库-资料库", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/知识库_资料库.png", doc,
                                  "知识库_资料库")

        # 作物知识库-知识浏览
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='作物知识库']").click()
        self.driver.find_element(By.XPATH,
                                 "//span[text()='作物知识库']/parent::div/following-sibling::ul//li[text()='知识浏览 ']").click()
        try:
            time.sleep(6)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@role='radiogroup']//span[text()='粮食作物']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开知识库-作物知识库-知识浏览页，且指定元素存在")
            image_elements = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//div[@id='tupianshow']//img")))
            unittest.TestCase.assertTrue(image_elements is not None,
                                         "成功打开知识库-作物知识库-知识浏览页，且指定元素存在")
            utils.g_logger.info("知识库-作物知识库-知识浏览页显示正常")
            sheet.append(["知识库-作物知识库-知识浏览", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/作物知识库_知识浏览.png", doc, "作物知识库_知识浏览")
        except Exception as e:
            utils.g_logger.info("知识库-作物知识库-知识浏览页异常")
            sheet.append(["知识库-作物知识库-知识浏览", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/作物知识库_知识浏览.png", doc, "作物知识库_知识浏览")

        # 作物知识库-知识维护
        time.sleep(3)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='作物知识库']/parent::div/following-sibling::ul//li[text()='知识维护 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='中文名']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开知识库-作物知识库-知识维护页，且指定元素存在")
            utils.g_logger.info("知识库-作物知识库-知识维护页显示正常")
            sheet.append(["知识库-作物知识库-知识维护", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/作物知识库_知识维护.png", doc, "作物知识库_知识维护")
        except Exception as e:
            utils.g_logger.info("知识库-作物知识库-知识维护页异常")
            sheet.append(["知识库-作物知识库-知识维护", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/作物知识库_知识维护.png", doc, "作物知识库_知识维护")

        # 关闭所有页面
        self.driver.find_element(By.XPATH, "//i[@class='el-icon-circle-close']").click()
        try:
            element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, "//ul[@class='el-dropdown-menu el-popper']/li[text()='关闭所有']")))
            element.click()
            utils.g_logger.info("成功关闭所有页面")
        except Exception as e:
            utils.g_logger.error(e)

        # 办公应用-工作平台页
        self.driver.find_element(By.XPATH, "//div[starts-with(@class,'navi-item')][text()='办公应用 ']").click()
        try:
            time.sleep(12)
            elements = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH,
                     "//div[@class='title'][text()='新闻管理']/parent::div/following-sibling::div/div[@class='content']")))
            unittest.TestCase.assertTrue(elements is not None, "成功打开办公应用-工作平台页，且指定元素存在")
            utils.g_logger.info("办公应用-工作平台页显示正常")
            sheet.append(["办公应用-工作平台", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_工作平台.png", doc, "办公应用_工作平台")
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["办公应用-工作平台", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_工作平台.png", doc, "办公应用_工作平台")

        # 办公应用-文件收发管理-收件箱页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='文件收发管理']").click()
        self.driver.find_element(By.XPATH,
                                 "//span[text()='文件收发管理']/parent::div/following-sibling::ul//li[text()='收件箱 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='收件时间']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-文件收发管理-收件箱页，且指定元素存在")
            utils.g_logger.info("办公应用-文件收发管理-收件箱页显示正常")
            sheet.append(["办公应用-文件收发管理-收件箱", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_文件收发管理_收件箱.png", doc,
                                  "办公应用_文件收发管理_收件箱")
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["办公应用-文件收发管理-收件箱", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_文件收发管理_收件箱.png", doc,
                                  "办公应用_文件收发管理_收件箱")

        # 办公应用-文件收发管理-草稿箱页
        time.sleep(3)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='文件收发管理']/parent::div/following-sibling::ul//li[text()='草稿箱 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='收件单位']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-文件收发管理-草稿箱页，且指定元素存在")
            utils.g_logger.info("办公应用-文件收发管理-草稿箱页显示正常")
            sheet.append(["办公应用-文件收发管理-草稿箱", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_文件收发管理_草稿箱.png", doc,
                                  "办公应用_文件收发管理_草稿箱")
        except Exception as e:
            utils.g_logger.info("办公应用-文件收发管理-草稿箱页异常")
            sheet.append(["办公应用-文件收发管理-草稿箱", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_文件收发管理_草稿箱.png", doc,
                                  "办公应用_文件收发管理_草稿箱")

        # 办公应用-文件收发管理-发件箱页
        time.sleep(3)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='文件收发管理']/parent::div/following-sibling::ul//li[text()='发件箱 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='发送时间']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-文件收发管理-发件箱页，且指定元素存在")
            utils.g_logger.info("办公应用-文件收发管理-发件箱页显示正常")
            sheet.append(["办公应用-文件收发管理-发件箱", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_文件收发管理_发件箱.png", doc,
                                  "办公应用_文件收发管理_发件箱")
        except Exception as e:
            utils.g_logger.info("办公应用-文件收发管理-发件箱页异常")
            sheet.append(["办公应用-文件收发管理-发件箱", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_文件收发管理_发件箱.png", doc,
                                  "办公应用_文件收发管理_发件箱")

        # 病虫害情报-情报管理页
        time.sleep(3)
        self.driver.find_element(By.XPATH, "//span[text()='病虫害情报']").click()
        self.driver.find_element(By.XPATH,
                                 "//span[text()='病虫害情报']/parent::div/following-sibling::ul//li[text()='情报管理 ']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//label[text()='情报类型']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-病虫害情报-情报管理页，且指定元素存在")
            utils.g_logger.info("办公应用-病虫害情报-情报管理页显示正常")
            sheet.append(["办公应用-病虫害情报-情报管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_病虫害情报_情报管理.png", doc,
                                  "办公应用_病虫害情报_情报管理")
        except Exception as e:
            utils.g_logger.info("办公应用-病虫害情报-情报管理页异常")
            sheet.append(["办公应用-病虫害情报-情报管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_病虫害情报_情报管理.png", doc,
                                  "办公应用_病虫害情报_情报管理")

        # 病虫害情报-情报库检索页
        time.sleep(5)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='病虫害情报']/parent::div/following-sibling::ul//li[text()='情报库检索 ']").click()
        try:
            time.sleep(15)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//label[text()='关键词']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-病虫害情报-情报库检索页，且指定元素存在")
            utils.g_logger.info("办公应用-病虫害情报-情报库检索页显示正常")
            sheet.append(["办公应用-病虫害情报-情报库检索", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_病虫害情报_情报库检索.png", doc,
                                  "办公应用_病虫害情报_情报库检索")
        except Exception as e:
            utils.g_logger.info("办公应用-病虫害情报-情报库检索页异常")
            sheet.append(["办公应用-病虫害情报-情报库检索", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_病虫害情报_情报库检索.png", doc,
                                  "办公应用_病虫害情报_情报库检索")

        # 病虫害情报-情报统计页
        time.sleep(5)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='病虫害情报']/parent::div/following-sibling::ul//li[text()='情报统计 ']").click()
        try:
            time.sleep(10)
            elements = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_any_elements_located(
                    (By.XPATH, "//span[@class='el-radio-button__inner'][contains(text(),'统计')]")))
            unittest.TestCase.assertTrue(elements is not None, "成功打开办公应用-病虫害情报-情报统计页，且指定元素存在")
            utils.g_logger.info("办公应用-病虫害情报-情报统计页显示正常")
            sheet.append(["办公应用-病虫害情报-情报统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_病虫害情报_情报统计.png", doc,
                                  "办公应用_病虫害情报_情报统计")
        except Exception as e:
            utils.g_logger.info("办公应用-病虫害情报-情报统计页异常")
            sheet.append(["办公应用-病虫害情报-情报统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_病虫害情报_情报统计.png", doc,
                                  "办公应用_病虫害情报_情报统计")

        # 新闻管理-新闻浏览页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//span[text()='新闻管理']/following-sibling::i").click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='新闻浏览 ']").click()
        try:
            time.sleep(10)
            elements = WebDriverWait(self.driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//h2[@class='zbtitle']")))
            unittest.TestCase.assertTrue(elements is not None, "成功打开办公应用-新闻管理-新闻浏览页，且指定元素存在")
            utils.g_logger.info("办公应用-新闻管理-新闻浏览页显示正常")
            sheet.append(["办公应用-新闻管理-新闻浏览", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_新闻管理_新闻浏览.png", doc,
                                  "办公应用_新闻管理_新闻浏览")
        except Exception as e:
            utils.g_logger.info("办公应用-新闻管理-新闻浏览页异常")
            sheet.append(["办公应用-新闻管理-新闻浏览", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_新闻管理_新闻浏览.png", doc,
                                  "办公应用_新闻管理_新闻浏览")

        # 新闻管理-新闻上传页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='新闻上传 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@class='cell'][text()='标题']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-新闻管理-新闻浏览页，且指定元素存在")
            utils.g_logger.info("办公应用-新闻管理-新闻上传页显示正常")
            sheet.append(["办公应用-新闻管理-新闻上传", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_新闻管理_新闻上传.png", doc,
                                  "办公应用_新闻管理_新闻上传")
        except Exception as e:
            utils.g_logger.info("办公应用-新闻管理-新闻上传页异常")
            sheet.append(["办公应用-新闻管理-新闻上传", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_新闻管理_新闻上传.png", doc,
                                  "办公应用_新闻管理_新闻上传")

        # 视频会议-会议管理页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//span[text()='视频会议']").click()
        time.sleep(2)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='会议管理 ']").click()
        try:
            time.sleep(8)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='参会人员']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-视频会议-会议管理页，且指定元素存在")
            utils.g_logger.info("办公应用-视频会议-会议管理页显示正常")
            sheet.append(["办公应用-视频会议-会议管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_视频会议_会议管理.png", doc,
                                  "办公应用_视频会议_会议管理")
        except Exception as e:
            utils.g_logger.info("办公应用-视频会议-会议管理页异常")
            sheet.append(["办公应用-视频会议-会议管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_视频会议_会议管理.png", doc,
                                  "办公应用_视频会议_会议管理")

        # 视频会议-我的会议页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//li[@role='menuitem'][text()='我的会议 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//label[@class='el-form-item__label'][text()='会议状态']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-视频会议-我的会议页，且指定元素存在")
            utils.g_logger.info("办公应用-视频会议-我的会议页显示正常")
            sheet.append(["办公应用-视频会议-我的会议", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_视频会议_我的会议.png", doc,
                                  "办公应用_视频会议_我的会议")
        except Exception as e:
            utils.g_logger.info("办公应用-视频会议-我的会议页异常")
            sheet.append(["办公应用-视频会议-我的会议", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_视频会议_我的会议.png", doc,
                                  "办公应用_视频会议_我的会议")

        # 通知公告-公告管理页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//span[text()='通知公告']").click()
        time.sleep(2)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='通知公告']/parent::div/following-sibling::ul//li[text()='公告管理 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='公告名称']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-通知公告-公告管理页，且指定元素存在")
            utils.g_logger.info("办公应用-通知公告-公告管理页显示正常")
            sheet.append(["办公应用-通知公告-公告管理", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_通知公告_公告管理.png", doc,
                                  "办公应用_通知公告_公告管理")
        except Exception as e:
            utils.g_logger.info("办公应用-通知公告-公告管理页异常")
            sheet.append(["办公应用-通知公告-公告管理", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_通知公告_公告管理.png", doc,
                                  "办公应用_通知公告_公告管理")

        # 通知公告-公告查阅页
        time.sleep(5)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='通知公告']/parent::div/following-sibling::ul//li[text()='公告查阅 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='公告名称']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-通知公告-公告查阅页，且指定元素存在")
            utils.g_logger.info("办公应用-通知公告-公告查阅页显示正常")
            sheet.append(["办公应用-通知公告-公告查阅", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_通知公告_公告查阅.png", doc,
                                  "办公应用_通知公告_公告查阅")
        except Exception as e:
            utils.g_logger.info("办公应用-通知公告-公告查阅页异常")
            sheet.append(["办公应用-通知公告-公告查阅", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_通知公告_公告查阅.png", doc,
                                  "办公应用_通知公告_公告查阅")

        # 办公应用-工作总结页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//span[text()='工作总结']").click()
        try:
            time.sleep(8)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//div[@class='cell'][text()='总结标题']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-工作总结页，且指定元素存在")
            utils.g_logger.info("办公应用-工作总结页显示正常")
            sheet.append(["办公应用-工作总结", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_工作总结.png", doc, "办公应用_工作总结")
        except Exception as e:
            utils.g_logger.info("办公应用-工作总结页异常")
            sheet.append(["办公应用-工作总结", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_工作总结.png", doc, "办公应用_工作总结")

        # 办公应用-业务考核-报送统计页
        time.sleep(5)
        self.driver.find_element(By.XPATH, "//span[text()='业务考核']").click()
        time.sleep(2)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='业务考核']/parent::div/following-sibling::ul//li[text()='报送统计 ']").click()
        try:
            time.sleep(10)
            element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH,
                                                  "//th[contains(@class,'     el-table__cell')]//div[@class='cell'][text()='《情报》统计']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-业务考核-报送统计页，且指定元素存在")
            utils.g_logger.info("办公应用-业务考核-报送统计页显示正常")
            sheet.append(["办公应用-业务考核-报送统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_业务考核_报送统计.png", doc,
                                  "办公应用_业务考核_报送统计")
        except Exception as e:
            utils.g_logger.info("办公应用-业务考核-报送统计页异常")
            sheet.append(["办公应用-业务考核-报送统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_业务考核_报送统计.png", doc,
                                  "办公应用_业务考核_报送统计")

        # 办公应用-业务考核-考核统计页
        time.sleep(5)
        self.driver.find_element(By.XPATH,
                                 "//span[text()='业务考核']/parent::div/following-sibling::ul//li[text()='考核统计 ']").click()
        try:
            time.sleep(12)
            element = WebDriverWait(self.driver, 12, 4).until(
                EC.visibility_of_element_located((By.XPATH,
                                                  "//th[contains(@class,'     is-leaf el-table__cell')]//div[@class='cell'][text()='总分']")))
            unittest.TestCase.assertTrue(element is not None, "成功打开办公应用-业务考核-考核统计页，且指定元素存在")
            utils.g_logger.info("办公应用-业务考核-考核统计页显示正常")
            sheet.append(["办公应用-业务考核-考核统计", "正常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_业务考核_考核统计.png", doc,
                                  "办公应用_业务考核_考核统计")
        except Exception as e:
            utils.g_logger.info("办公应用-业务考核-考核统计页异常")
            sheet.append(["办公应用-业务考核-考核统计", "异常"])
            utils.page_screenshot(self.driver, "outputs/imagefiles/办公应用_业务考核_考核统计.png", doc,
                                  "办公应用_业务考核_考核统计")
        doc.save('outputs/广西截图.docx')

        # 调整列宽
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 8)
            sheet.column_dimensions[column].width = adjusted_width

    # 广西草地贪夜蛾综合管理平台
    @utils.retry(MAX_TRIES)
    def test_guangXiTye(self):
        # 创建一个sheet并加上名称和所在位置，第一个位置索引号是0
        wb.create_sheet("广西草地贪夜蛾综合管理平台", 2)
        sheet = wb["广西草地贪夜蛾综合管理平台"]
        # 写入表头
        headers = ["页面", "检测结果"]
        sheet.append(headers)
        username = input("广西草贪系统生产环境巡检开始\n请输入登录用户名：")
        password = input("请输入登录密码：")
        '''guangXi_cfg = cfg["guangXi"]
        base_url = guangXi_cfg["baseUrl"]
        username = guangXi_cfg["username"]
        password = guangXi_cfg["password"]'''
        self.driver.maximize_window()
        # self.driver.get(base_url + "login")
        self.driver.get("https://gxtye.pestiot.com")
        self.driver.implicitly_wait(4)

        # 登录
        self.driver.find_element(By.XPATH, '//input[@placeholder="请输入用户名"]').send_keys(username)
        self.driver.find_element(By.XPATH, '//input[@type="password"]').send_keys(password)
        yanzhengma_image = self.driver.find_element(By.XPATH, '//img[contains(@src,"/ppms/sys/user/getvalidatecode")]')
        img_bytes = yanzhengma_image.screenshot_as_png
        yzm = ddddocr.DdddOcr(show_ad=False).classification(img_bytes)
        self.driver.find_element(By.XPATH, '//input[@placeholder="请输入验证码"]').send_keys(yzm)
        time.sleep(3)
        self.driver.find_element(By.XPATH, '//button[@type="button"]//span[text()="登录"]').click()

        # 首页全年任务数量元素校验
        try:
            element = WebDriverWait(self.driver, 15).until(EC.visibility_of_element_located(
                (By.XPATH,
                 '//div[@class="ivu-col ivu-col-span-xs-24 ivu-col-span-sm-12 ivu-col-span-md-6"][1]')))  # 定位全年任务数量
            unittest.TestCase.assertTrue(element is not None, "登录成功，成功打开首页，且指定元素存在")
            utils.g_logger.info(f"登录成功，成功打开首页")
            sheet.append(["首页", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["首页", "异常"])

        # 项目管理-立项管理-项目组管理页
        self.driver.find_element(By.XPATH, "//ul[contains(@class,'navli ivu-menu')]//li[2]").click()  # 项目管理
        try:
            time.sleep(3)
            element = WebDriverWait(self.driver, 15).until(EC.visibility_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(element is not None, "成功打开项目管理-立项管理-项目组管理，且指定元素存在")
            utils.g_logger.info("项目管理-立项管理-项目组管理页显示正常")
            sheet.append(["项目管理-立项管理-项目组管理", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-立项管理-项目组管理", "异常"])

        # 项目管理-立项管理-项目档案页
        self.driver.find_element(By.XPATH,
                                 "//*[i[@data-v-36cafb69=''] and contains(text(), '立项管理')]").click()  # 项目管理-立项管理
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'ivu-menu-item') and text()='项目档案']").click()  # 立项管理-项目档案

        try:
            time.sleep(3)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None,
                                         "成功打开项目管理-立项管理-项目档案页，且指定元素存在")
            utils.g_logger.info("项目管理-立项管理-项目档案页显示正常")
            sheet.append(["项目管理-立项管理-项目档案", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-立项管理-项目档案", "异常"])

        # 项目管理-立项管理-立项审核页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'ivu-menu-item') and text()='立项审核']").click()  # 项目管理-立项审核

        try:
            time.sleep(3)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None,
                                         "成功打开项目管理-立项管理-项目档案页，且指定元素存在")
            utils.g_logger.info("项目管理-立项管理-立项审核页显示正常")
            sheet.append(["项目管理-立项管理-立项审核", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-立项管理-立项审核", "异常"])

        Xpath_Road_Start = ["//*[i[@data-v-36cafb69=''] and contains(text(), '招投标管理')]",
                            "//*[i[@data-v-36cafb69=''] and contains(text(), '合同管理')]",
                            "//*[i[@data-v-36cafb69=''] and contains(text(), '资金管理')]",
                            "//*[i[@data-v-36cafb69=''] and contains(text(), '建设进度')]"
                            ]
        # 项目管理-招投标管理页
        self.driver.find_element(By.XPATH,
                                 Xpath_Road_Start[0]).click()  # 招投标管理
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开项目管理-招投标管理页")
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='建设单位']/following-sibling::div/div").click()  # 点击建设单位选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='dropdown']")))  # 建设单位站点
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开项目管理-招投标管理页，且站点列表存在")
            utils.g_logger.info("项目管理-招投标管理页显示正常")
            sheet.append(["项目管理-招投标管理", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-招投标管理", "异常"])

        # 项目管理-合同管理页
        self.driver.find_element(By.XPATH,
                                 Xpath_Road_Start[1]).click()  # 合同管理
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开项目管理-招投标管理页")
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='建设单位']/following-sibling::div/div").click()  # 点击建设单位选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='dropdown']")))  # 建设单位站点
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开项目管理-合同管理页，且站点列表存在")
            utils.g_logger.info("项目管理-合同管理页显示正常")
            sheet.append(["项目管理-合同管理", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-合同管理", "异常"])

        # 项目管理-资金管理页
        self.driver.find_element(By.XPATH,
                                 Xpath_Road_Start[2]).click()  # 资金管理
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开项目管理-招投标管理页")
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='建设单位']/following-sibling::div/div").click()  # 点击建设单位选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='dropdown']")))  # 建设单位站点
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开项目管理-资金管理页，且站点列表存在")
            utils.g_logger.info("项目管理-资金管理页显示正常")
            sheet.append(["项目管理-资金管理", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-资金管理", "异常"])

        # 项目管理-建设进度页
        self.driver.find_element(By.XPATH,
                                 Xpath_Road_Start[3]).click()  # 建设进度
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开项目管理-招投标管理页")
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='建设单位']/following-sibling::div/div").click()  # 点击建设单位选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='dropdown']")))  # 建设单位站点
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开项目管理-建设进度页，且站点列表存在")
            utils.g_logger.info("项目管理-建设进度页显示正常")
            sheet.append(["项目管理-建设进度", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-建设进度", "异常"])

        # 验收管理-验收申请页
        self.driver.find_element(By.XPATH,
                                 "//*[i[@data-v-36cafb69=''] and contains(text(), '验收管理')]").click()  # 验收管理
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'ivu-menu-item') and text()='验收申请']").click()  # 验收申请
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开验收管理-验收申请页")
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='建设单位']/following-sibling::div/div").click()  # 点击建设单位选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='dropdown']")))  # 建设单位站点
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开验收管理-验收申请页，且站点列表存在")
            utils.g_logger.info("验收管理-验收申请显示正常")
            sheet.append(["验收管理-验收申请", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["验收管理-验收申请", "异常"])

        # 验收管理-验收审核页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'ivu-menu-item') and text()='验收审核']").click()  # 验收审核
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[contains(@class,"ivu-input") and @placeholder="起始时间"]')))  # 立项时间-起始时间
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开验收管理-验收审核页")
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='建设单位']/following-sibling::div/div").click()  # 点击建设单位选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='dropdown']")))  # 建设单位站点
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开验收管理-验收审核页，且站点列表存在")
            utils.g_logger.info("验收管理-验收审核显示正常")
            sheet.append(["验收管理-验收审核", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["验收管理-验收审核", "异常"])

        # 项目管理-绩效考核页
        self.driver.find_element(By.XPATH,
                                 "//*[i[@data-v-36cafb69=''] and contains(text(), '绩效考核')]").click()  # 绩效考核
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[contains(@class,'ivu-select ivu-select-single')]//span[text()='请选择']")))  # 项目组
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开项目管理-绩效考核页")
            self.driver.find_element(By.XPATH,
                                     "//input[@class='ivu-input']").click()  # 点击年度选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='ivu-picker-panel-body']")))  # 年度选择框年度列表
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开项目管理-绩效考核页，且年度列表存在")
            utils.g_logger.info("项目管理-绩效考核页显示正常")
            sheet.append(["项目管理-绩效考核", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-绩效考核", "异常"])

        # 项目管理-数据统计页
        self.driver.find_element(By.XPATH,
                                 "//*[i[@data-v-36cafb69=''] and contains(text(), '数据统计')]").click()  # 数据统计
        self.driver.find_element(By.XPATH,
                                 "//li[@class='ivu-menu-item' and contains(text(), '数据统计')]").click()  # 数据统计，不稳定，点击两次
        try:
            time.sleep(5)
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[contains(@class,'ivu-select ivu-select-single')]//span[text()='请选择']")))  # 项目组
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开项目管理-数据统计页")
            self.driver.find_element(By.XPATH,
                                     "//input[@class='ivu-input']").click()  # 点击年度选择框
            orgnization_tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='ivu-picker-panel-body']")))  # 年度选择框年度列表
            unittest.TestCase.assertTrue(orgnization_tree_element is not None,
                                         "成功打开项目管理-数据统计页，且年度列表存在")
            utils.g_logger.info("项目管理-数据统计页显示正常")
            sheet.append(["项目管理-数据统计", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["项目管理-数据统计", "异常"])

        # 数据填报-数据填报页
        self.driver.find_element(By.XPATH, "//ul[contains(@class,'navli ivu-menu')]//li[3]").click()
        self.driver.find_element(By.XPATH, "//li[@class='ivu-menu-item' and contains(text(),'数据填报')]").click()
        try:
            time.sleep(3)
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='报表']/following-sibling::div/div").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[@class="ivu-input" and @placeholder="报表分类"]')))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开数据填报页，且指定元素存在")
            utils.g_logger.info("数据填报-数据填报页显示正常")
            sheet.append(["数据填报-数据填报", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["数据填报-数据填报", "异常"])

        # 数据填报数据查询页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'ivu-menu-item') and contains(text(),'数据查询')]").click()
        try:
            time.sleep(3)
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='报表']/following-sibling::div/div").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[@class="ivu-input" and @placeholder="报表分类"]')))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开数据查询页，且指定元素存在")
            utils.g_logger.info("数据填报-数据查询页显示正常")
            sheet.append(["数据填报-数据查询", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["数据填报-数据查询", "异常"])

        # 数据填报-数据汇总页
        self.driver.find_element(By.XPATH,
                                 "//li[contains(@class,'ivu-menu-item') and contains(text(),'数据汇总')]").click()
        try:
            time.sleep(3)
            self.driver.find_element(By.XPATH,
                                     "//label[@class='ivu-form-item-label' and text()='报表']/following-sibling::div/div").click()
            tree_element = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located(
                (By.XPATH, '//input[@class="ivu-input" and @placeholder="报表分类"]')))
            unittest.TestCase.assertTrue(tree_element is not None, "成功打开数据汇总页，且指定元素存在")
            utils.g_logger.info("数据填报-数据汇总页显示正常")
            sheet.append(["数据填报-数据汇总", "正常"])
        except Exception as e:
            utils.g_logger.info(e)
            sheet.append(["数据填报-数据汇总", "异常"])

        # 调整列宽
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 8)
            sheet.column_dimensions[column].width = adjusted_width

    def export_excel(self):
        wb.save('outputs/省级功能巡检.xlsx')
        utils.g_logger.info("省级系统功能巡检结束，请检查输出的巡检文件内容。")


if __name__ == '__main__':
    p = PPMSGX()
    # 广西省级系统页面巡检
    p.test_guangXi()
    # 将所有巡检结果导出excel文件
    p.export_excel()
